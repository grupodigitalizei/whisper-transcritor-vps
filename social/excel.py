"""Exportação Excel/CSV das coletas. Portado de ~/IGSorter/igsorter/excel.py.

Adaptações: usa social.core.EXPORT_DIR/build_trends e degrada com elegância quando
o Pillow não está instalado — nesse caso o Excel sai sem as miniaturas (o resto,
incluindo a aba de Tendências, continua igual). O CSV nunca depende de nada extra.
"""
import csv
import datetime as dt
import io
import os
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .core import EXPORT_DIR, build_trends

try:
    from PIL import Image as PILImage
    _PIL_OK = True
except Exception:
    _PIL_OK = False


def thumbs_supported() -> bool:
    return _PIL_OK


def _fetch_thumb(url, size=(76, 76)):
    if not _PIL_OK:
        return None
    try:
        import requests
        r = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        im = PILImage.open(io.BytesIO(r.content)).convert("RGB")
        im.thumbnail(size)
        buf = io.BytesIO()
        im.save(buf, "PNG")
        return buf.getvalue()
    except Exception:
        return None


def _header(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def export_excel(dataset, sort="er", thumbs=True, out=None, csv_too=True,
                 on_progress=None):
    thumbs = thumbs and _PIL_OK
    rows = list(dataset["rows"])
    keymap = {"er": lambda r: r["er"] if r["er"] is not None else -1,
              "views": lambda r: r["views"] or -1,
              "likes": lambda r: r["likes"],
              "comments": lambda r: r["comments"],
              "date": lambda r: r["ts"] or 0}
    rows.sort(key=keymap.get(sort, keymap["er"]), reverse=True)

    username = (dataset.get("profile") or {}).get("username") or \
        (rows[0]["username"] if rows else "perfil")
    out = out or os.path.join(EXPORT_DIR, f"{username}_{dt.date.today()}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"
    headers = ["Thumb", "Data", "Tipo", "Views", "Likes", "Comentários", "Reshares",
               "ER %", "Duração (s)", "Link", "Legenda"]
    _header(ws, headers)
    for i, w in enumerate([13, 17, 12, 12, 10, 12, 10, 8, 11, 34, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, r in enumerate(rows, start=2):
        ws.row_dimensions[i].height = 60 if thumbs else 16
        if thumbs and r["thumb_url"]:
            data = _fetch_thumb(r["thumb_url"])
            if data:
                try:
                    from openpyxl.drawing.image import Image as XLImage
                    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                    tmp.write(data)
                    tmp.close()
                    ws.add_image(XLImage(tmp.name), f"A{i}")
                except Exception:
                    pass
        if on_progress and i % 10 == 0:
            on_progress(i - 1)
        date = dt.datetime.fromisoformat(r["date"]) if r["date"] else None
        ws.cell(i, 2, date.strftime("%d/%m/%Y %H:%M") if date else "")
        ws.cell(i, 3, r["type"])
        ws.cell(i, 4, r["views"])
        ws.cell(i, 5, r["likes"])
        ws.cell(i, 6, r["comments"])
        ws.cell(i, 7, r["reshares"])
        ws.cell(i, 8, r["er"])
        ws.cell(i, 9, r["duration_s"])
        link = ws.cell(i, 10, r["url"])
        link.hyperlink = r["url"]
        link.font = Font(color="0563C1", underline="single")
        ws.cell(i, 11, (r["caption"] or "")[:250])
    ws.auto_filter.ref = f"A1:K{len(rows) + 1}"

    trends = build_trends(rows)
    wt = wb.create_sheet("Tendências")
    row_i = 1
    for title, key in [("Por dia da semana", "weekday"), ("Por hora do dia", "hour"),
                       ("Por formato", "type"), ("Por duração (reels)", "duration"),
                       ("Top hashtags (min. 2 posts)", "hashtag")]:
        wt.cell(row_i, 1, title).font = Font(bold=True, size=12, color="1F4E79")
        row_i += 1
        for col, h in enumerate(["", "Posts", "ER médio %", "Views médias", "Likes médios"], 1):
            wt.cell(row_i, col, h).font = Font(bold=True)
        row_i += 1
        for t in trends[key]:
            wt.cell(row_i, 1, str(t["key"]))
            wt.cell(row_i, 2, t["posts"])
            wt.cell(row_i, 3, t.get("er_medio"))
            wt.cell(row_i, 4, t.get("views_medias"))
            wt.cell(row_i, 5, t.get("likes_medios"))
            row_i += 1
        row_i += 1
    for i, w in enumerate([26, 8, 12, 14, 14], 1):
        wt.column_dimensions[get_column_letter(i)].width = w

    profile = dataset.get("profile")
    if profile:
        wp = wb.create_sheet("Perfil")
        for j, (k, v) in enumerate(profile.items(), 1):
            wp.cell(j, 1, k).font = Font(bold=True)
            wp.cell(j, 2, str(v))
        wp.column_dimensions["A"].width = 18
        wp.column_dimensions["B"].width = 60

    wb.save(out)
    result = {"excel": out}

    if csv_too:
        csv_path = out.rsplit(".", 1)[0] + ".csv"
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["data", "tipo", "views", "likes", "comentarios", "reshares",
                        "er_pct", "duracao_s", "link", "legenda"])
            for r in rows:
                w.writerow([r["date"] or "", r["type"], r["views"], r["likes"],
                            r["comments"], r["reshares"], r["er"], r["duration_s"],
                            r["url"], (r["caption"] or "").replace("\n", " ")[:300]])
        result["csv"] = csv_path
    return result
